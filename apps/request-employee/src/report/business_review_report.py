from __future__ import annotations
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

def write_business_review_report(
    output_path: str | Path,
    request_no: str,
    pdf_doc,
    excel_doc,
    diffs: list[dict],
    reconciliation: dict,
    action: str,
    acceptance: dict,
):
    wb = Workbook()

    ws = wb.active
    ws.title = "業務確認"
    ws["A1"] = "請求書 照合結果"
    ws["A1"].font = Font(bold=True, size=16)
    ws.append([])
    rows = [
        ("請求書番号", request_no),
        ("PDF顧客名", pdf_doc.customer_name),
        ("Excel顧客名", excel_doc.customer_name),
        ("PDF請求日", pdf_doc.request_date),
        ("Excel請求日", excel_doc.request_date),
        ("PDF合計", str(pdf_doc.total_amount)),
        ("Excel合計", str(excel_doc.total_amount)),
        ("差異件数", len(diffs)),
        ("判定", action),
        ("受入評価", acceptance.get("grade")),
        ("受入点数", acceptance.get("score")),
    ]
    for k, v in rows:
        ws.append([k, v])

    for cell in ws["A"]:
        cell.font = Font(bold=True)
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 50

    ws2 = wb.create_sheet("差異一覧")
    ws2.append(["区分", "項目", "PDF", "Excel", "重要度", "差異種別"])
    for cell in ws2[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9EAF7")
    for d in diffs:
        ws2.append([
            d.get("scope"),
            d.get("field"),
            d.get("pdf"),
            d.get("excel"),
            d.get("severity"),
            d.get("difference_type", ""),
        ])
    ws2.freeze_panes = "A2"
    ws2.auto_filter.ref = ws2.dimensions

    ws3 = wb.create_sheet("整合性チェック")
    ws3.append(["対象", "ルール", "期待値", "実際値", "重要度", "メッセージ"])
    for cell in ws3[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="FCE4D6")
    for target in ("pdf", "excel"):
        for item in reconciliation.get(target, []):
            ws3.append([
                target.upper(),
                item.get("rule"),
                item.get("expected"),
                item.get("actual"),
                item.get("severity"),
                item.get("message"),
            ])

    ws4 = wb.create_sheet("PDF明細")
    ws4.append(["No", "商品コード", "商品名", "数量", "単価", "金額", "税率"])
    for line in pdf_doc.lines:
        ws4.append([line.line_no, line.product_code, line.product_name, float(line.quantity), float(line.unit_price), float(line.amount), float(line.tax_rate)])

    ws5 = wb.create_sheet("Excel明細")
    ws5.append(["No", "商品コード", "商品名", "数量", "単価", "金額", "税率"])
    for line in excel_doc.lines:
        ws5.append([line.line_no, line.product_code, line.product_name, float(line.quantity), float(line.unit_price), float(line.amount), float(line.tax_rate)])

    for sheet in wb.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        for col in sheet.columns:
            max_len = max(len(str(c.value or "")) for c in col)
            sheet.column_dimensions[col[0].column_letter].width = min(max(max_len + 2, 12), 45)

    wb.save(output_path)
