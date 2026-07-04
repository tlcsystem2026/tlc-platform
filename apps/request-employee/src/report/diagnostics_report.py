from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

def write_diagnostics_workbook(rows: list[dict], output_path: str | Path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Parser Diagnostics"
    headers = [
        "Source",
        "Request No",
        "Request Date",
        "Customer",
        "Total",
        "Line Count",
        "Missing Fields",
        "Status",
    ]
    ws.append(headers)
    for c in ws[1]:
        c.font = Font(bold=True)

    fill_bad = PatternFill("solid", fgColor="FFC7CE")
    fill_ok = PatternFill("solid", fgColor="C6EFCE")

    for row in rows:
        ws.append([
            row.get("source_file"),
            row.get("request_no"),
            row.get("request_date"),
            row.get("customer_name"),
            row.get("total_amount"),
            row.get("line_count"),
            ", ".join(row.get("missing_fields", [])),
            row.get("status"),
        ])
        fill = fill_ok if row.get("status") == "OK" else fill_bad
        for cell in ws[ws.max_row]:
            cell.fill = fill

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for col in ws.columns:
        max_len = max(len(str(c.value or "")) for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(max(max_len + 2, 12), 50)
    wb.save(output_path)
