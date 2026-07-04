from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

SEVERITY_FILL = {
    "ERROR": PatternFill("solid", fgColor="FFC7CE"),
    "WARNING": PatternFill("solid", fgColor="FFEB9C"),
    "INFO": PatternFill("solid", fgColor="D9EAF7"),
}

def write_difference_report(diffs: list[dict], output_path: str | Path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Differences"
    headers = ["Scope", "Field", "PDF", "Excel", "Severity", "Status"]
    ws.append(headers)
    for c in ws[1]:
        c.font = Font(bold=True)

    for d in diffs:
        ws.append([
            d.get("scope"),
            d.get("field"),
            d.get("pdf"),
            d.get("excel"),
            d.get("severity", "INFO"),
            d.get("status", "DIFFERENT"),
        ])
        fill = SEVERITY_FILL.get(d.get("severity", "INFO"))
        if fill:
            for cell in ws[ws.max_row]:
                cell.fill = fill

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for col in ws.columns:
        max_len = max(len(str(c.value or "")) for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(max(max_len + 2, 12), 40)
    wb.save(output_path)
