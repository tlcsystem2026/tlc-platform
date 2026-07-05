from __future__ import annotations
from pathlib import Path
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font

HEADERS = [
    "登録日時", "請求書番号", "請求日", "顧客名",
    "商品コード", "商品名", "数量", "単価", "金額", "税率",
    "PDFファイル", "Excelファイル",
]

class SalesLedger:
    def __init__(self, path: str | Path):
        self.path = Path(path)
        self.path.parent.mkdir(parents=True, exist_ok=True)

    def append_request(self, request_doc, pdf_path: str = "", excel_path: str = "") -> int:
        wb, ws = self._open()
        now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        before = ws.max_row

        if not request_doc.lines:
            ws.append([
                now, request_doc.request_no, request_doc.request_date, request_doc.customer_name,
                "", "", "", "", float(request_doc.total_amount), "", pdf_path, excel_path
            ])
        else:
            for line in request_doc.lines:
                ws.append([
                    now, request_doc.request_no, request_doc.request_date, request_doc.customer_name,
                    line.product_code, line.product_name, float(line.quantity), float(line.unit_price),
                    float(line.amount), float(line.tax_rate), pdf_path, excel_path
                ])

        self._format(ws)
        wb.save(self.path)
        return ws.max_row - before

    def _open(self):
        if self.path.exists():
            wb = load_workbook(self.path)
            ws = wb.active
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = "Sales Ledger"
            ws.append(HEADERS)
            for cell in ws[1]:
                cell.font = Font(bold=True)
            ws.freeze_panes = "A2"
        return wb, ws

    def _format(self, ws):
        widths = {"A":20,"B":16,"C":14,"D":28,"E":18,"F":36,"G":10,"H":12,"I":14,"J":10,"K":45,"L":45}
        for col, width in widths.items():
            ws.column_dimensions[col].width = width
        ws.auto_filter.ref = ws.dimensions
