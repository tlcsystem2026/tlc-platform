from __future__ import annotations
from pathlib import Path
import csv
import pdfplumber
from openpyxl import load_workbook

def export_pdf_raw(pdf_path: str | Path, output_path: str | Path):
    chunks = []
    with pdfplumber.open(str(pdf_path)) as pdf:
        for page_no, page in enumerate(pdf.pages, 1):
            chunks.append(f"===== PAGE {page_no} =====")
            chunks.append(page.extract_text(x_tolerance=2, y_tolerance=3) or "")
    Path(output_path).write_text("\n".join(chunks), encoding="utf-8")

def export_excel_cells(excel_path: str | Path, output_path: str | Path):
    wb = load_workbook(excel_path, data_only=False, read_only=False)
    with Path(output_path).open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(["sheet", "cell", "value", "data_type"])
        for ws in wb.worksheets:
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value is not None:
                        writer.writerow([ws.title, cell.coordinate, cell.value, cell.data_type])
