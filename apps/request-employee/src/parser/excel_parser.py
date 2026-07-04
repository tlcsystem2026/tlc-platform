from __future__ import annotations
from pathlib import Path
from openpyxl import load_workbook

from models.request import RequestDocument, RequestLine
from parser.normalizer import clean_text, normalize_key, normalize_money

ALIASES = {
    "request_no": ["請求書番号", "請求番号", "requestno", "no"],
    "request_date": ["請求日", "請求年月日", "日付"],
    "customer_name": ["顧客名", "得意先名", "請求先", "会社名"],
    "product_code": ["商品コード", "品番", "コード"],
    "product_name": ["商品名", "品名", "摘要"],
    "quantity": ["数量", "個数"],
    "unit_price": ["単価"],
    "amount": ["金額"],
}

class ExcelParser:
    def parse(self, path: str | Path) -> RequestDocument:
        path = Path(path)
        wb = load_workbook(path, data_only=True, read_only=False)
        ws = wb.active
        doc = RequestDocument(source_file=str(path))
        cells = list(ws.iter_rows(values_only=True))
        doc.request_no = self._find_label_value(cells, ALIASES["request_no"])
        doc.request_date = clean_text(self._find_label_value(cells, ALIASES["request_date"]))
        doc.customer_name = self._find_label_value(cells, ALIASES["customer_name"])
        doc.lines = self._parse_table(cells)
        doc.total_amount = sum((x.amount for x in doc.lines), normalize_money(0))
        return doc

    def _find_label_value(self, rows, aliases):
        wanted = {normalize_key(x) for x in aliases}
        for row in rows:
            for i, value in enumerate(row):
                if normalize_key(value) in wanted:
                    for j in range(i + 1, len(row)):
                        if row[j] not in (None, ""):
                            return clean_text(row[j])
        return ""

    def _parse_table(self, rows):
        header_row = None
        mapping = {}
        for r_idx, row in enumerate(rows):
            normalized = [normalize_key(x) for x in row]
            local = {}
            for field in ("product_code", "product_name", "quantity", "unit_price", "amount"):
                aliases = {normalize_key(x) for x in ALIASES[field]}
                for c_idx, value in enumerate(normalized):
                    if value in aliases:
                        local[field] = c_idx
                        break
            if "product_name" in local and "amount" in local:
                header_row, mapping = r_idx, local
                break
        if header_row is None:
            return []
        result = []
        for row in rows[header_row + 1:]:
            def get(field):
                i = mapping.get(field)
                return row[i] if i is not None and i < len(row) else ""
            name = clean_text(get("product_name"))
            if not name or any(x in name for x in ["小計", "合計", "消費税"]):
                continue
            result.append(RequestLine(
                line_no=len(result) + 1,
                product_code=clean_text(get("product_code")),
                product_name=name,
                quantity=normalize_money(get("quantity")),
                unit_price=normalize_money(get("unit_price")),
                amount=normalize_money(get("amount")),
            ))
        return result
