from __future__ import annotations
from pathlib import Path
from decimal import Decimal
import re
from datetime import datetime, date
from openpyxl import load_workbook

from models.request import RequestDocument, RequestLine
from parser.normalizer import clean_text, normalize_money

class TokyoKoibitoExcelParser:
    """
    Tokyo Koibito Excel parser.

    Build017 improvements:
    - reads known layout directly
    - extracts customer/date/request_no/total
    - extracts tax-excluded subtotal and tax values
    - ignores blank placeholder rows
    """

    def parse(self, path: str | Path) -> RequestDocument:
        path = Path(path)
        wb = load_workbook(path, data_only=True)
        ws = wb.active
        cells = [[c.value for c in row] for row in ws.iter_rows()]
        flat = [clean_text(v) for row in cells for v in row if v not in (None, "")]

        doc = RequestDocument(source_file=str(path))
        doc.request_no = self._request_no(flat)
        doc.request_date = self._request_date(cells)
        doc.customer_name = self._customer(cells)
        doc.total_amount = self._total(cells)
        doc.subtotal, doc.tax_amount = self._subtotal_tax(cells)
        doc.lines = self._lines(cells)

        if doc.total_amount == 0:
            doc.total_amount = sum((x.amount for x in doc.lines), Decimal("0"))
        return doc

    def _request_no(self, flat: list[str]) -> str:
        for v in flat:
            m = re.search(r"(LY\d{4,})", v)
            if m:
                return m.group(1)
        return ""

    def _request_date(self, cells) -> str:
        for row in cells:
            for i, v in enumerate(row):
                if clean_text(v) == "請求日":
                    for j in range(i + 1, len(row)):
                        d = row[j]
                        if isinstance(d, (datetime, date)):
                            return f"{d.year:04d}-{d.month:02d}-{d.day:02d}"
                        s = clean_text(d)
                        m = re.search(r"(\d{4})年(\d{1,2})月(\d{1,2})日", s)
                        if m:
                            y, mo, day = map(int, m.groups())
                            return f"{y:04d}-{mo:02d}-{day:02d}"
        return ""

    def _customer(self, cells) -> str:
        for row in cells:
            vals = [clean_text(v) for v in row if v not in (None, "")]
            if len(vals) >= 2 and vals[-1] == "御中":
                return vals[0]
        return ""

    def _total(self, cells):
        total = Decimal("0")
        for row in cells:
            vals = [v for v in row if v not in (None, "")]
            if len(vals) >= 2 and "ご請求額" in clean_text(vals[0]):
                total = normalize_money(vals[-1])
        return total

    def _subtotal_tax(self, cells):
        subtotal = Decimal("0")
        tax = Decimal("0")
        for row in cells:
            vals = [v for v in row if v not in (None, "")]
            if len(vals) >= 2 and clean_text(vals[-2]) == "小計（税抜）":
                subtotal += normalize_money(vals[-1])
            if len(vals) >= 3 and clean_text(vals[0]) == "消費税":
                tax += normalize_money(vals[-1])
        return subtotal, tax

    def _lines(self, cells) -> list[RequestLine]:
        results = []
        tax_rate = Decimal("0.10")
        in_table = False

        for row in cells:
            row_text = " ".join(clean_text(v) for v in row if v not in (None, ""))
            if "②軽減税率商品" in row_text:
                tax_rate = Decimal("0.08")
                in_table = False
                continue
            if "①通常税率商品" in row_text:
                tax_rate = Decimal("0.10")
                in_table = False
                continue
            if "番号" in row_text and "商品名" in row_text and "金額" in row_text:
                in_table = True
                continue
            if not in_table:
                continue

            vals = [v for v in row if v not in (None, "")]
            if len(vals) < 5:
                continue
            if not str(vals[0]).strip().isdigit():
                continue

            line_no = int(vals[0])
            name = clean_text(vals[1])
            qty = normalize_money(vals[2])
            unit_price = normalize_money(vals[3])
            amount = normalize_money(vals[4])

            if not name or amount == 0:
                continue

            product_code = ""
            for token in name.split():
                if re.search(r"\d{8,}", token):
                    product_code = token
                    break

            results.append(RequestLine(
                line_no=line_no,
                product_code=product_code,
                product_name=name,
                quantity=qty,
                unit_price=unit_price,
                amount=amount,
                tax_rate=tax_rate,
            ))

        return results
