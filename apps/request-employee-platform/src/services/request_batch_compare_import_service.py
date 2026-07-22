from __future__ import annotations

import csv
import hashlib
import json
import re
import shutil
import unicodedata
from datetime import datetime, timezone
from decimal import Decimal, InvalidOperation
from pathlib import Path
from uuid import uuid4

from openpyxl import load_workbook
from sqlalchemy import text
from sqlalchemy.orm import Session

from src.services.request_folder_settings_service import ensure_month_directories, validate_business_month
from src.services.request_tax_breakdown_service import (
    BREAKDOWN_FIELDS,
    compare_tax_breakdowns,
    extract_tax_breakdown_from_excel,
    extract_tax_breakdown_from_text,
)

try:
    from pypdf import PdfReader
except Exception:
    try:
        from PyPDF2 import PdfReader
    except Exception:
        PdfReader = None

BATCH_TABLE = "tlc_request_batch_compare"
ITEM_TABLE = "tlc_request_batch_compare_item"
REVIEW_TABLE = "tlc_request_review_queue"
PDF_EXTENSIONS = {".pdf"}
EXCEL_EXTENSIONS = {".xlsx", ".xlsm", ".xls"}


def ensure_tables(db: Session) -> None:
    db.execute(text(f"""CREATE TABLE IF NOT EXISTS {BATCH_TABLE}(
        id VARCHAR(64) PRIMARY KEY,business_month VARCHAR(6) NOT NULL,
        status VARCHAR(32) NOT NULL,operator VARCHAR(255) NOT NULL,
        source_directory TEXT NOT NULL,total_file_count INTEGER NOT NULL DEFAULT 0,
        pair_count INTEGER NOT NULL DEFAULT 0,review_count INTEGER NOT NULL DEFAULT 0,
        exception_count INTEGER NOT NULL DEFAULT 0,error_count INTEGER NOT NULL DEFAULT 0,
        exception_report_path TEXT NOT NULL DEFAULT '',started_at VARCHAR(64) NOT NULL,
        completed_at VARCHAR(64) NOT NULL DEFAULT '',message TEXT NOT NULL DEFAULT '')"""))
    db.execute(text(f"""CREATE TABLE IF NOT EXISTS {ITEM_TABLE}(
        id VARCHAR(64) PRIMARY KEY,batch_id VARCHAR(64) NOT NULL,business_month VARCHAR(6) NOT NULL,
        pair_key VARCHAR(500) NOT NULL,pdf_file_name VARCHAR(500) NOT NULL DEFAULT '',
        excel_file_name VARCHAR(500) NOT NULL DEFAULT '',original_pdf_path TEXT NOT NULL DEFAULT '',
        original_excel_path TEXT NOT NULL DEFAULT '',final_pdf_path TEXT NOT NULL DEFAULT '',
        final_excel_path TEXT NOT NULL DEFAULT '',pdf_sha256 VARCHAR(64) NOT NULL DEFAULT '',
        excel_sha256 VARCHAR(64) NOT NULL DEFAULT '',pdf_raw_text TEXT NOT NULL DEFAULT '',
        excel_raw_json TEXT NOT NULL DEFAULT '',raw_customer_name VARCHAR(500) NOT NULL DEFAULT '',
        system_customer_code VARCHAR(255) NOT NULL DEFAULT '',system_customer_name VARCHAR(500) NOT NULL DEFAULT '',
        customer_match_status VARCHAR(32) NOT NULL DEFAULT 'UNMATCHED',compare_status VARCHAR(32) NOT NULL,
        exception_codes TEXT NOT NULL DEFAULT '',exception_details TEXT NOT NULL DEFAULT '',
        pdf_total_amount VARCHAR(64) NOT NULL DEFAULT '',excel_total_amount VARCHAR(64) NOT NULL DEFAULT '',
        review_status VARCHAR(32) NOT NULL DEFAULT 'WAIT_REVIEW',created_at VARCHAR(64) NOT NULL)"""))
    db.execute(text(f"""CREATE TABLE IF NOT EXISTS {REVIEW_TABLE}(
        id VARCHAR(64) PRIMARY KEY,batch_id VARCHAR(64) NOT NULL,item_id VARCHAR(64) NOT NULL UNIQUE,
        business_month VARCHAR(6) NOT NULL,pair_key VARCHAR(500) NOT NULL,
        review_status VARCHAR(32) NOT NULL DEFAULT 'WAIT_REVIEW',compare_status VARCHAR(32) NOT NULL,
        raw_customer_name VARCHAR(500) NOT NULL DEFAULT '',system_customer_code VARCHAR(255) NOT NULL DEFAULT '',
        system_customer_name VARCHAR(500) NOT NULL DEFAULT '',exception_codes TEXT NOT NULL DEFAULT '',
        created_at VARCHAR(64) NOT NULL)"""))

    item_columns = {
        row[1]
        for row in db.execute(
            text(f"PRAGMA table_info({ITEM_TABLE})")
        ).all()
    }
    tax_columns = {
        "pdf_tax_breakdown_json": "TEXT NOT NULL DEFAULT '{}'",
        "excel_tax_breakdown_json": "TEXT NOT NULL DEFAULT '{}'",
    }
    for side in ("pdf", "excel"):
        for field in BREAKDOWN_FIELDS:
            tax_columns[f"{side}_{field}"] = (
                "VARCHAR(64) NOT NULL DEFAULT ''"
            )
    for column, definition in tax_columns.items():
        if column not in item_columns:
            db.execute(
                text(
                    f"ALTER TABLE {ITEM_TABLE} "
                    f"ADD COLUMN {column} {definition}"
                )
            )
    # BUILD037_BATCH_CURRENT_FLAG_SCHEMA_R6
    batch_columns = {
        row[1]
        for row in db.execute(
            text(f"PRAGMA table_info({BATCH_TABLE})")
        ).all()
    }
    if "is_current" not in batch_columns:
        db.execute(
            text(
                f"ALTER TABLE {BATCH_TABLE} "
                "ADD COLUMN is_current INTEGER NOT NULL DEFAULT 0"
            )
        )

    review_columns = {
        row[1]
        for row in db.execute(
            text(f"PRAGMA table_info({REVIEW_TABLE})")
        ).all()
    }
    if "is_current" not in review_columns:
        db.execute(
            text(
                f"ALTER TABLE {REVIEW_TABLE} "
                "ADD COLUMN is_current INTEGER NOT NULL DEFAULT 0"
            )
        )
    db.commit()


def _now() -> str:
    return datetime.now(timezone.utc).isoformat()


def _sha256(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def _pair_key(path: Path) -> str:
    value = path.stem.lower().strip()
    for suffix in ("_pdf", "-pdf", " pdf", "_excel", "-excel", " excel", "_xlsx", "-xlsx", " xlsx", "_請求書", "-請求書", " 請求書"):
        if value.endswith(suffix):
            value = value[:-len(suffix)].rstrip(" _-")
    return re.sub(r"\s+", "", value)


def _extract_pdf(path: Path) -> str:
    if PdfReader is None:
        raise RuntimeError("PDF parser is not installed")
    return "\n".join((page.extract_text() or "") for page in PdfReader(str(path)).pages).strip()


def _extract_excel(path: Path) -> dict:
    if path.suffix.lower() == ".xls":
        raise RuntimeError("Legacy .xls is not supported")
    workbook = load_workbook(path, data_only=True, read_only=True)
    result = {"sheets": []}
    for sheet in workbook.worksheets:
        rows = []
        for row in sheet.iter_rows(values_only=True):
            values = [v.isoformat() if hasattr(v, "isoformat") else v for v in row]
            if any(v not in (None, "") for v in values):
                rows.append(values)
        result["sheets"].append({"title": sheet.title, "rows": rows})
    workbook.close()
    return result


def _flatten_excel(data: dict) -> str:
    return "\n".join(str(v) for sheet in data.get("sheets", []) for row in sheet.get("rows", []) for v in row if v not in (None, ""))


def _amounts(value: str) -> list[Decimal]:
    result = []
    pattern = r"(?<!\d)(?:¥|￥)?\s*([0-9]{1,3}(?:,[0-9]{3})+|[0-9]+)(?:\.\d{1,2})?(?!\d)"
    for matched in re.findall(pattern, value or ""):
        try:
            result.append(Decimal(matched.replace(",", "")))
        except InvalidOperation:
            pass
    return result


def _amount_string(value: Decimal) -> str:
    return str(value.quantize(Decimal("0.01"))).rstrip("0").rstrip(".")


def _likely_total(value: str) -> str:
    values = _amounts(value)
    if not values:
        return ""
    return _amount_string(max(values))


def _label_key(value: object) -> str:
    return re.sub(r"[\\s　・･,，。．.\-_/\\()（）「」『』①②＋+]", "", str(value or "").lower())


def _pdf_labeled_total(value: str) -> str:
    text_value = unicodedata.normalize("NFKC", str(value or ""))
    labels = (
        r"ご\s*請\s*求\s*額",
        r"御\s*請\s*求\s*金\s*額",
        r"御\s*請\s*求\s*額",
        r"合\s*計\s*請\s*求\s*額",
        r"請\s*求\s*金\s*額",
    )
    candidates: list[Decimal] = []
    amount_pattern = r"([0-9]{1,3}(?:,[0-9]{3})+|[0-9]+)(?:\.\d{1,2})?"

    for label in labels:
        for label_match in re.finditer(label, text_value, flags=re.IGNORECASE):
            suffix = text_value[label_match.end() : label_match.end() + 120]

            currency_match = re.search(
                rf"(?:¥|￥|\\)\s*{amount_pattern}",
                suffix,
            )
            if currency_match:
                try:
                    candidates.append(
                        Decimal(currency_match.group(1).replace(",", ""))
                    )
                    continue
                except InvalidOperation:
                    pass

            comma_match = re.search(
                rf"(?<![0-9]){amount_pattern}(?![0-9])",
                suffix,
            )
            if comma_match:
                try:
                    candidates.append(
                        Decimal(comma_match.group(1).replace(",", ""))
                    )
                except InvalidOperation:
                    pass

    if candidates:
        return _amount_string(candidates[-1])
    return _likely_total(text_value)


def _excel_labeled_total(data: dict) -> str:
    labels = ("ご請求額", "合計請求額", "請求金額")
    candidates: list[Decimal] = []
    for sheet in data.get("sheets", []):
        for row in sheet.get("rows", []):
            for index, cell in enumerate(row):
                key = _label_key(cell)
                if not any(_label_key(label) in key for label in labels):
                    continue
                for value in row[index + 1 :]:
                    if value in (None, ""):
                        continue
                    if isinstance(value, (int, float, Decimal)):
                        try:
                            candidates.append(Decimal(str(value)))
                            break
                        except InvalidOperation:
                            continue
                    amounts = _amounts(str(value))
                    if amounts:
                        candidates.append(amounts[-1])
                        break
    if candidates:
        return _amount_string(candidates[-1])
    return _likely_total(_flatten_excel(data))


def _clean_customer_candidate(value: object) -> str:
    candidate = unicodedata.normalize(
        "NFKC",
        str(value or ""),
    ).strip()
    candidate = re.sub(
        r"\s*(?:御中|様|殿)\s*$",
        "",
        candidate,
    ).strip()
    return candidate


def _likely_customer(data: dict) -> str:
    # The invoice recipient is the name immediately before or beside
    # honorifics such as 御中. This must take priority over the issuer name.
    for sheet in data.get("sheets", []):
        rows = sheet.get("rows", [])[:40]
        for row_index, row in enumerate(rows):
            for cell_index, value in enumerate(row):
                cell = unicodedata.normalize(
                    "NFKC",
                    str(value or ""),
                ).strip()
                if cell not in {"御中", "様", "殿"}:
                    continue

                # Same row, immediately to the left.
                for previous in reversed(row[:cell_index]):
                    candidate = _clean_customer_candidate(previous)
                    if candidate:
                        return candidate

                # Previous non-empty row, which is common in invoice forms.
                for previous_row in reversed(rows[:row_index]):
                    for previous in previous_row:
                        candidate = _clean_customer_candidate(previous)
                        if candidate:
                            return candidate

    # Fallback for older layouts without an explicit honorific cell.
    issuer_tokens = {
        "東京恋人株式会社",
        "トウキヨウコイビトカブシキガイシヤ",
    }
    for sheet in data.get("sheets", []):
        for row in sheet.get("rows", [])[:30]:
            for value in row[:12]:
                candidate = _clean_customer_candidate(value)
                if not candidate or candidate in issuer_tokens:
                    continue
                if (
                    2 <= len(candidate) <= 120
                    and any(
                        token in candidate
                        for token in (
                            "株式会社",
                            "有限会社",
                            "合同会社",
                            "法人",
                            "会社",
                            "商事",
                            "ストア",
                            "ショップ",
                            "店",
                        )
                    )
                ):
                    return candidate
    return ""


def _normalize(value: str) -> str:
    normalized = unicodedata.normalize("NFKC", str(value or "")).casefold()
    normalized = re.sub(
        r"[\s\u3000・･·•∙⋅,，。．.\-‐‑‒–—―_/\\()（）「」『』【】]",
        "",
        normalized,
    )
    replacements = {
        "㈱": "株式会社",
        "(株)": "株式会社",
        "（株）": "株式会社",
        "㈲": "有限会社",
        "(有)": "有限会社",
        "（有）": "有限会社",
    }
    for source, target in replacements.items():
        normalized = normalized.replace(source, target)
    return normalized


def _pdf_recipient_name(pdf_text: str) -> str:
    normalized = unicodedata.normalize("NFKC", str(pdf_text or ""))
    for match in re.finditer(
        r"(?m)^\s*(.+?)\s*(?:御中|様|殿)\s*$",
        normalized,
    ):
        candidate = _clean_customer_candidate(match.group(1))
        if candidate and candidate != "東京恋人株式会社":
            return candidate
    lines = [line.strip() for line in normalized.splitlines() if line.strip()]
    for index, line in enumerate(lines):
        if line in {"御中", "様", "殿"} and index > 0:
            candidate = _clean_customer_candidate(lines[index - 1])
            if candidate and candidate != "東京恋人株式会社":
                return candidate
    return ""


def _customer_name_found_in_pdf(raw_customer: str, pdf_text: str) -> bool:
    raw_normalized = _normalize(raw_customer)
    if not raw_normalized:
        return False
    recipient = _pdf_recipient_name(pdf_text)
    if recipient:
        return raw_normalized == _normalize(recipient)
    return raw_normalized in _normalize(pdf_text)


def _match_customer(db: Session, raw_name: str) -> tuple[str, str, str]:
    if not raw_name:
        return "", "", "UNMATCHED"

    normalized_raw = _normalize(raw_name)

    for table_name in (
        "tlc_customer",
        "tlc_customer_master",
        "tlc_customers",
    ):
        exists = db.execute(
            text(
                "SELECT name FROM sqlite_master "
                "WHERE type='table' AND name=:name"
            ),
            {"name": table_name},
        ).first()
        if not exists:
            continue

        if table_name == "tlc_customer_master":
            try:
                columns = {
                    str(row._mapping["name"])
                    for row in db.execute(
                        text(
                            "PRAGMA table_info(tlc_customer_master)"
                        )
                    ).all()
                }
                candidate_columns = [
                    column
                    for column in (
                        "formal_name",
                        "hiragana_name",
                        "katakana_name",
                        "katakana_name_short",
                        "short_name",
                        "delivery_name_1",
                        "delivery_name_2",
                        "alias_1",
                        "alias_2",
                        "alias_3",
                        "alias_4",
                        "alias_5",
                    )
                    if column in columns
                ]
                select_columns = ",".join(candidate_columns)
                rows = db.execute(
                    text(
                        "SELECT customer_id,"
                        + select_columns
                        + " FROM tlc_customer_master"
                    )
                ).all()
                for row in rows:
                    mapping = row._mapping
                    for column in candidate_columns:
                        value = str(mapping.get(column) or "")
                        if value and _normalize(value) == normalized_raw:
                            formal_name = str(
                                mapping.get("formal_name")
                                or value
                            )
                            return (
                                str(mapping.get("customer_id") or ""),
                                formal_name,
                                "MATCHED",
                            )
            except Exception:
                pass

        for code_col, name_col in (
            ("customer_id", "formal_name"),
            ("id", "formal_name"),
            ("customer_id", "customer_name"),
        ):
            try:
                rows = db.execute(
                    text(
                        f"SELECT {code_col} code,"
                        f"{name_col} name FROM {table_name}"
                    )
                ).all()
            except Exception:
                continue
            for row in rows:
                if _normalize(row._mapping["name"]) == normalized_raw:
                    return (
                        str(row._mapping["code"] or ""),
                        str(row._mapping["name"] or ""),
                        "MATCHED",
                    )
    return "", "", "UNMATCHED"


def _move(source: Path, destination: Path) -> Path:
    destination.parent.mkdir(parents=True, exist_ok=True)
    if destination.exists():
        destination = destination.with_name(f"{destination.stem}_{uuid4().hex[:8]}{destination.suffix}")
    shutil.move(str(source), str(destination))
    return destination


def latest_batch(db: Session, business_month: str = "") -> dict | None:
    ensure_tables(db)
    params = {}
    where = ""
    if business_month:
        params["month"] = validate_business_month(business_month)
        where = "WHERE business_month=:month"
    row = db.execute(text(f"SELECT * FROM {BATCH_TABLE} {where} ORDER BY started_at DESC LIMIT 1"), params).first()
    return dict(row._mapping) if row else None


def list_review_queue(db: Session, business_month: str = "", limit: int = 1000) -> list[dict]:
    ensure_tables(db)
    params = {"limit": min(max(int(limit), 1), 2000)}
    where = ""
    if business_month:
        params["month"] = validate_business_month(business_month)
        where = "WHERE business_month=:month"
    rows = db.execute(text(f"SELECT * FROM {REVIEW_TABLE} {where} ORDER BY created_at DESC LIMIT :limit"), params).all()
    return [dict(row._mapping) for row in rows]


def run_request_batch(db: Session, *, business_month: str, operator: str) -> dict:
    ensure_tables(db)
    month = validate_business_month(business_month)
    operator = str(operator or "").strip()
    if not operator:
        raise ValueError("operator is required")
    dirs = ensure_month_directories(month)
    incoming, processing = Path(dirs["incoming"]), Path(dirs["processing"])
    completed, error = Path(dirs["completed"]), Path(dirs["error"])
    files = sorted(p for p in incoming.iterdir() if p.is_file() and p.suffix.lower() in (PDF_EXTENSIONS | EXCEL_EXTENSIONS))
    batch_id = uuid4().hex
    run_date = datetime.now().strftime("%Y%m%d")
    error_run_dir = error / run_date / batch_id
    db.execute(text(f"INSERT INTO {BATCH_TABLE}(id,business_month,status,operator,source_directory,total_file_count,started_at) VALUES(:id,:month,'PROCESSING',:operator,:source,:count,:started_at)"), {"id": batch_id, "month": month, "operator": operator, "source": str(incoming), "count": len(files), "started_at": _now()})
    db.commit()
    groups = {}
    for source in files:
        moved = _move(source, processing / source.name)
        groups.setdefault(_pair_key(moved), {"pdf": [], "excel": []})["pdf" if moved.suffix.lower() in PDF_EXTENSIONS else "excel"].append(moved)
    exception_rows = []
    pair_count = review_count = error_count = 0
    for pair_key, group in groups.items():
        for index in range(max(len(group["pdf"]), len(group["excel"]))):
            pdf_path = group["pdf"][index] if index < len(group["pdf"]) else None
            excel_path = group["excel"][index] if index < len(group["excel"]) else None
            codes, details = [], []
            pdf_text, excel_data = "", {"sheets": []}
            pdf_total = excel_total = raw_customer = ""
            pdf_tax_breakdown = {}
            excel_tax_breakdown = {}
            fatal = False
            if pdf_path is None:
                codes.append("PDF_MISSING"); details.append("Matching PDF file is missing")
            if excel_path is None:
                codes.append("EXCEL_MISSING"); details.append("Matching Excel file is missing")
            try:
                if pdf_path: pdf_text = _extract_pdf(pdf_path)
                if excel_path: excel_data = _extract_excel(excel_path)
                if pdf_path and excel_path:
                    excel_text = _flatten_excel(excel_data)
                    pdf_total, excel_total = _pdf_labeled_total(pdf_text), _excel_labeled_total(excel_data)
                    pdf_tax_breakdown = extract_tax_breakdown_from_text(pdf_text)
                    excel_tax_breakdown = extract_tax_breakdown_from_excel(excel_data)
                    tax_codes, tax_details = compare_tax_breakdowns(
                        pdf_tax_breakdown,
                        excel_tax_breakdown,
                    )
                    codes.extend(tax_codes)
                    details.extend(tax_details)
                    raw_customer = _likely_customer(excel_data)
                    if not pdf_text: codes.append("PDF_TEXT_EMPTY"); details.append("PDF text could not be extracted")
                    if not excel_text: codes.append("EXCEL_DATA_EMPTY"); details.append("Excel contains no readable data")
                    if pdf_total and excel_total and pdf_total != excel_total:
                        codes.append("TOTAL_AMOUNT_MISMATCH"); details.append(f"PDF total={pdf_total}, Excel total={excel_total}")
                    if raw_customer and not _customer_name_found_in_pdf(raw_customer, pdf_text):
                        codes.append("CUSTOMER_NAME_NOT_FOUND_IN_PDF")
                        details.append(
                            "Excel customer name does not match PDF recipient: "
                            f"{raw_customer}"
                        )
                compare_status = "MATCHED" if not codes else "EXCEPTION"
            except Exception as exc:
                fatal = True; compare_status = "ERROR"; codes.append("PROCESSING_ERROR"); details.append(str(exc)); error_count += 1
            customer_code, customer_name, customer_status = _match_customer(db, raw_customer)
            destination = completed if compare_status == "MATCHED" else error_run_dir
            final_pdf = _move(pdf_path, destination / pdf_path.name) if pdf_path and pdf_path.exists() else None
            final_excel = _move(excel_path, destination / excel_path.name) if excel_path and excel_path.exists() else None
            item_id, created_at = uuid4().hex, _now()
            db.execute(text(f"""INSERT INTO {ITEM_TABLE}(id,batch_id,business_month,pair_key,pdf_file_name,excel_file_name,original_pdf_path,original_excel_path,final_pdf_path,final_excel_path,pdf_sha256,excel_sha256,pdf_raw_text,excel_raw_json,raw_customer_name,system_customer_code,system_customer_name,customer_match_status,compare_status,exception_codes,exception_details,pdf_total_amount,excel_total_amount,review_status,created_at) VALUES(:id,:batch_id,:month,:pair_key,:pdf_name,:excel_name,:original_pdf,:original_excel,:final_pdf,:final_excel,:pdf_hash,:excel_hash,:pdf_text,:excel_json,:raw_customer,:customer_code,:customer_name,:customer_status,:compare_status,:codes,:details,:pdf_total,:excel_total,'WAIT_REVIEW',:created_at)"""), {"id": item_id, "batch_id": batch_id, "month": month, "pair_key": pair_key, "pdf_name": pdf_path.name if pdf_path else "", "excel_name": excel_path.name if excel_path else "", "original_pdf": str(incoming / pdf_path.name) if pdf_path else "", "original_excel": str(incoming / excel_path.name) if excel_path else "", "final_pdf": str(final_pdf or ""), "final_excel": str(final_excel or ""), "pdf_hash": _sha256(final_pdf) if final_pdf else "", "excel_hash": _sha256(final_excel) if final_excel else "", "pdf_text": pdf_text, "excel_json": json.dumps(excel_data, ensure_ascii=False, default=str), "raw_customer": raw_customer, "customer_code": customer_code, "customer_name": customer_name, "customer_status": customer_status, "compare_status": compare_status, "codes": ",".join(codes), "details": "\n".join(details), "pdf_total": pdf_total, "excel_total": excel_total, "created_at": created_at})
            tax_params = {
                "id": item_id,
                "pdf_tax_breakdown_json": json.dumps(
                    pdf_tax_breakdown,
                    ensure_ascii=False,
                ),
                "excel_tax_breakdown_json": json.dumps(
                    excel_tax_breakdown,
                    ensure_ascii=False,
                ),
            }
            tax_assignments = [
                "pdf_tax_breakdown_json=:pdf_tax_breakdown_json",
                "excel_tax_breakdown_json=:excel_tax_breakdown_json",
            ]
            for side, breakdown in (
                ("pdf", pdf_tax_breakdown),
                ("excel", excel_tax_breakdown),
            ):
                for field in BREAKDOWN_FIELDS:
                    key = f"{side}_{field}"
                    tax_params[key] = str(breakdown.get(field, "") or "")
                    tax_assignments.append(f"{key}=:{key}")
            db.execute(
                text(
                    f"UPDATE {ITEM_TABLE} SET "
                    + ",".join(tax_assignments)
                    + " WHERE id=:id"
                ),
                tax_params,
            )
            db.execute(text(f"INSERT INTO {REVIEW_TABLE}(id,batch_id,item_id,business_month,pair_key,compare_status,raw_customer_name,system_customer_code,system_customer_name,exception_codes,created_at,is_current) VALUES(:id,:batch_id,:item_id,:month,:pair_key,:compare_status,:raw_customer,:customer_code,:customer_name,:codes,:created_at,0)"), {"id": uuid4().hex, "batch_id": batch_id, "item_id": item_id, "month": month, "pair_key": pair_key, "compare_status": compare_status, "raw_customer": raw_customer, "customer_code": customer_code, "customer_name": customer_name, "codes": ",".join(codes), "created_at": created_at})
            db.commit(); pair_count += 1; review_count += 1
            if compare_status != "MATCHED":
                exception_rows.append({"batch_id": batch_id, "business_month": month, "pair_key": pair_key, "pdf_file_name": pdf_path.name if pdf_path else "", "excel_file_name": excel_path.name if excel_path else "", "compare_status": compare_status, "exception_codes": ",".join(codes), "exception_details": " | ".join(details), "raw_customer_name": raw_customer, "system_customer_code": customer_code, "system_customer_name": customer_name})
    report_path = ""
    if exception_rows:
        report = error_run_dir / f"request_exceptions_{month}_{batch_id[:8]}.csv"
        with report.open("w", encoding="utf-8-sig", newline="") as handle:
            writer = csv.DictWriter(handle, fieldnames=list(exception_rows[0]))
            writer.writeheader(); writer.writerows(exception_rows)
        report_path = str(report)
    message = f"files={len(files)}, pairs={pair_count}, review={review_count}, exceptions={len(exception_rows)}, errors={error_count}"
    db.execute(text(f"UPDATE {BATCH_TABLE} SET status=:status,pair_count=:pairs,review_count=:reviews,exception_count=:exceptions,error_count=:errors,exception_report_path=:report,completed_at=:completed_at,message=:message WHERE id=:id"), {"status": "COMPLETED_WITH_ERRORS" if exception_rows else "COMPLETED", "pairs": pair_count, "reviews": review_count, "exceptions": len(exception_rows), "errors": error_count, "report": report_path, "completed_at": _now(), "message": message, "id": batch_id})
    # BUILD037_BATCH_PROMOTE_CURRENT_R6
    db.execute(
        text(
            f"UPDATE {BATCH_TABLE} SET is_current=0 "
            "WHERE business_month=:month"
        ),
        {"month": month},
    )
    db.execute(
        text(f"UPDATE {BATCH_TABLE} SET is_current=1 WHERE id=:id"),
        {"id": batch_id},
    )
    db.execute(
        text(
            f"UPDATE {REVIEW_TABLE} SET is_current=0 "
            "WHERE business_month=:month"
        ),
        {"month": month},
    )
    db.execute(
        text(
            f"UPDATE {REVIEW_TABLE} SET is_current=1 "
            "WHERE batch_id=:id"
        ),
        {"id": batch_id},
    )
    db.commit()
    return latest_batch(db, month)
